#!/usr/bin/env python3
"""
pharma-epi-qc — read-only standardization QC for the Epidemiology Master Sheet.

Checks the Granular Data tab against the Indication Aliases tab (source of truth
for canonical indication names + therapeutic areas) and the fixed 7-type metric
taxonomy. Emits a Markdown exceptions report. NEVER writes to the workbook —
fixes are applied separately, only after human sign-off.

Usage:
    python3 qc.py "Epidemiology Master Sheet (DD Month YYYY, HMMam).xlsx" \
        [--out qc_report.md]
"""
import argparse
import re
import sys
from collections import defaultdict

import openpyxl

# The valid metric taxonomy is NOT hardcoded — it is read from the workbook's
# "Metric Type Aliases" tab (col A) at runtime, so the taxonomy lives in exactly
# one place. An approved new type added to that tab flows to QC automatically.

# Granular Data column indices (1-based), data begins col A (no spacer).
COL = {
    "source": 1, "company": 2, "disease_area": 3, "orig_ind": 4, "std_ind": 5,
    "orig_metric": 6, "metric_type": 7, "us": 8, "eu5": 9, "china": 10,
    "japan": 11, "total": 12, "geography": 13, "year": 14, "method": 15, "dq": 16,
}

# Below this relative difference, a Total vs sum-of-geographies mismatch is
# treated as source rounding/float noise, not a real error.
ROUNDING_TOLERANCE = 0.005


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_aliases(ws):
    """Return (canon_ta, canon_by_norm, alias_to_canon) from the Aliases tab.

    A data row has both col A (canonical) and col B (TA). Header/group rows have
    col A but no col B and are skipped.
    """
    canon_ta = {}
    canon_by_norm = {}
    alias_to_canon = {}
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if a and b:
            canon_ta[a] = b
            canon_by_norm[norm(a)] = a
            alias_to_canon[norm(a)] = a
            if c:
                for tok in str(c).split(","):
                    tok = tok.strip()
                    if tok:
                        alias_to_canon.setdefault(norm(tok), a)
    return canon_ta, canon_by_norm, alias_to_canon


def load_metric_types(ws):
    """Valid metric taxonomy = col A of the 'Metric Type Aliases' tab. A data row
    has both col A (canonical type) and col B (definition); the header and HOW-TO
    rows have col A but no col B, so they are skipped."""
    types = set()
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a and b:
            types.add(str(a).strip())
    return types


def run(path, out):
    wb = openpyxl.load_workbook(path, data_only=True)
    for tab in ("Indication Aliases", "Granular Data", "Metric Type Aliases"):
        if tab not in wb.sheetnames:
            sys.exit(f"FAIL: workbook missing {tab!r} tab.")
    al = wb["Indication Aliases"]
    ws = wb["Granular Data"]
    canon_ta, canon_by_norm, alias_to_canon = load_aliases(al)
    metric_types = load_metric_types(wb["Metric Type Aliases"])
    if not metric_types:
        sys.exit("FAIL: 'Metric Type Aliases' tab has no taxonomy rows.")

    orphan = []          # col E matches nothing
    alias_not_canon = []  # col E is an alias, not the canonical
    ta_mismatch = []      # col C != canonical TA
    e_to_c = defaultdict(set)
    bad_metric = []       # col G not in the 7 types
    f_to_g = defaultdict(set)
    inferred = []
    red_rows = []         # material Total != sum
    rounding_rows = []    # sub-tolerance mismatch
    n = 0

    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, COL["source"]).value:
            continue
        n += 1
        E = ws.cell(r, COL["std_ind"]).value
        C = ws.cell(r, COL["disease_area"]).value
        G = ws.cell(r, COL["metric_type"]).value
        F = ws.cell(r, COL["orig_metric"]).value
        P = str(ws.cell(r, COL["dq"]).value or "")

        if E:
            nE = norm(E)
            if nE in canon_by_norm:
                cn = canon_by_norm[nE]
                e_to_c[cn].add(C)
                if C and canon_ta[cn] != C:
                    ta_mismatch.append((r, E, C, canon_ta[cn]))
            elif nE in alias_to_canon:
                alias_not_canon.append((r, E, alias_to_canon[nE]))
            else:
                orphan.append((r, E))

        if G and G not in metric_types:
            bad_metric.append((r, G))
        if F and G:
            f_to_g[str(F).strip()].add(G)
        if "inferred" in P.lower():
            inferred.append(r)

        geos = [to_num(ws.cell(r, COL[k]).value) for k in ("us", "eu5", "china", "japan")]
        present = [g for g in geos if g is not None]
        tot = to_num(ws.cell(r, COL["total"]).value)
        if len(present) >= 2 and tot is not None:
            s = sum(present)
            denom = max(abs(s), abs(tot)) or 1
            diff = abs(s - tot) / denom
            if diff > ROUNDING_TOLERANCE:
                red_rows.append((r, s, tot, diff))
            elif diff > 0:
                rounding_rows.append(r)

    multi_ta = {cn: {x for x in cs if x} for cn, cs in e_to_c.items()
                if len({x for x in cs if x}) > 1}
    metric_conflict = {f: gs for f, gs in f_to_g.items() if len(gs) > 1}

    L = []
    w = L.append
    w("# Epidemiology Master Sheet — Standardization QC Report\n")
    w(f"- **Workbook:** {path}")
    w(f"- **Data rows checked:** {n}")
    w(f"- **Canonicals in Aliases tab:** {len(canon_ta)}\n")
    w("Read-only diagnostic. No cells were changed. This is investment-research")
    w("material — verify before use; not a sole basis for a decision.\n")

    w("## Summary\n")
    w("| Check | Result |")
    w("|---|---|")
    w(f"| Indication names (col E) map to a canonical | {'PASS' if not orphan and not alias_not_canon else 'FAIL'} |")
    w(f"| Metric type (col G) in taxonomy ({len(metric_types)} types) | {'PASS' if not bad_metric else 'FAIL'} |")
    w(f"| Same metric label -> same type | {'PASS' if not metric_conflict else 'FAIL'} |")
    w(f"| Disease Area (col C) matches canonical TA | {'PASS' if not ta_mismatch else 'FAIL (%d rows)' % len(ta_mismatch)} |")
    w(f"| Metric type confidence | {len(inferred)} rows flagged 'inferred' |")
    w(f"| Total != sum (material, >{ROUNDING_TOLERANCE:.1%}) | {len(red_rows)} rows |")
    w(f"| Total != sum (rounding, ignored) | {len(rounding_rows)} rows |\n")

    w("## TIER 1 — Indication & Disease Area\n")
    w("**Orphan col E (no canonical/alias match):** " + (str([r for r, _ in orphan]) if orphan else "none"))
    for r, e in orphan:
        w(f"- r{r}: {e!r}")
    w("\n**col E is an alias, not the canonical (replace):** " + (str([r for r, _, _ in alias_not_canon]) if alias_not_canon else "none"))
    for r, e, cn in alias_not_canon:
        w(f"- r{r}: {e!r} -> {cn!r}")
    w("\n**Disease Area (col C) != canonical TA:**")
    if ta_mismatch:
        w("\n| row | indication | col C (wrong) | should be |")
        w("|---|---|---|---|")
        for r, e, c, ta in ta_mismatch:
            w(f"| {r} | {e} | {c} | {ta} |")
    else:
        w(" none")
    w("\n**Same indication used with multiple Disease Area labels:**")
    for cn, cs in multi_ta.items():
        w(f"- {cn!r}: {cs}")
    if not multi_ta:
        w("- none")

    w("\n## TIER 2 — Metric standardization\n")
    w("**col G outside the 7 taxonomy types:** " + (str(bad_metric) if bad_metric else "none — PASS"))
    w("\n**Same Original Metric Label -> multiple standardized types:**")
    for f, gs in metric_conflict.items():
        w(f"- {f[:60]!r}: {gs}")
    if not metric_conflict:
        w("- none — PASS")
    w(f"\n**'inferred — verify' rows ({len(inferred)}):** {inferred}")

    w("\n## TIER 3 — Total vs sum-of-geographies\n")
    w("**Material discrepancies (flag red):**")
    if red_rows:
        w("\n| row | sum geos | stated Total | diff % |")
        w("|---|---|---|---|")
        for r, s, tot, diff in red_rows:
            w(f"| {r} | {s:,.0f} | {tot:,.0f} | {diff*100:.1f}% |")
    else:
        w(" none")
    w(f"\n**Rounding-only mismatches left unflagged ({len(rounding_rows)}):** {rounding_rows}")

    with open(out, "w") as fh:
        fh.write("\n".join(L) + "\n")
    print(f"Wrote {out}: {len(orphan)} orphan, {len(alias_not_canon)} alias-not-canon, "
          f"{len(ta_mismatch)} TA mismatch, {len(bad_metric)} bad metric, "
          f"{len(metric_conflict)} metric conflicts, {len(inferred)} inferred, "
          f"{len(red_rows)} material Total!=sum, {len(rounding_rows)} rounding.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--out", default="qc_report.md")
    a = ap.parse_args()
    run(a.workbook, a.out)


if __name__ == "__main__":
    main()
