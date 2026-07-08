#!/usr/bin/env python3
"""pharma-epi-exporter — append extractor staging CSVs into a LOCAL copy of the
Epidemiology Master Sheet (.xlsx).

This never touches Box. It reads a local copy of the master workbook plus the
extractor's staging files and writes a new `*_UPDATED.xlsx` next to it, leaving
the source workbook untouched. The user reviews the updated file and uploads it
to Box as a new version themselves.

Usage:
    python3 export.py \
        --workbook "/path/Epidemiology Master Sheet.xlsx" \
        --granular  /path/<slug>_granular.csv \
        --sheet1    /path/<slug>_sheet1.csv \
        [--aliases  /path/<slug>_aliases.csv] \
        [--link-data URL] \
        [--out /path/out.xlsx] [--dry-run]
"""

import argparse
import csv
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Flag fills (two-color model). The extractor writes a token in the Data Quality
# Notes cell (col P); this script paints the fill in one pass.
#   flag:uncertain -> soft value/classification (averaged/midpoint/range/inferred
#                     /chart-read) -> FFD966 on the populated numeric cell(s)
#   flag:red       -> known-wrong (e.g. Total != sum of geographies) -> red on Total
UNCERTAIN_FILL = "FFD966"
RED_FILL = "FFFF0000"
TOTAL_COL = 12  # L
GRANULAR_SHEET = "Granular Data"
SHEET1_SHEET = "Sheet1"
ALIASES_SHEET = "Indication Aliases"

# Granular Data column layout (1-indexed): A..P = 16 columns, data starts col A.
GRANULAR_NCOLS = 16
NUMERIC_COLS = (8, 9, 10, 11, 12)  # H US, I EU5, J China, K Japan, L Total
NUM_FORMAT = "#,##0"

# Granular Data visual style: zebra banding (light blue / white) + Arial 10 wrap.
GRANULAR_BANDS = ("FFEBF3FB", "FFFFFFFF")  # even row -> light blue, odd -> white
GRANULAR_FONT = Font(name="Arial", size=10)
GRANULAR_ALIGN = Alignment(wrap_text=True, vertical="top")
GRANULAR_ROW_H = 28.0

# Sheet1: col A is a blank spacer; the 8 staging fields map to cols B..I
# (header is in row 2). There is no companion-doc column.
SHEET1_START_COL = 2      # B
SHEET1_END_COL = 9        # I
SHEET1_NFIELDS = 8
SHEET1_LINK_DATA_COL = 7  # G Link to data

# Indication Aliases: TA-grouped, color-coded. Group-header rows are dark blue,
# bold, white text; data rows take a group-specific pastel fill.
ALIAS_HEADER_FILL = "FF2E75B6"
ALIAS_NEWGROUP_FILL = "FFF2F2F2"  # placeholder fill for a brand-new TA group
ALIAS_ROW_H = 30.0
ALIAS_HEADER_H = 20.0


def fail(msg):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


# --- Enforced checkpoint gate (between extraction and export) -----------------
# A real write is refused unless a human has signed off the review note. The
# contract is deterministic, with no CLI override (that would defeat the gate):
#   1. The note must contain a CHECKED sign-off line:
#        - [x] Checkpoint: approved for export
#   2. No unresolved items may remain (any unchecked "- [ ]" box blocks).
# The extractor writes the note with the box UNCHECKED; the human checks it at
# the checkpoint after resolving open items. A --dry-run writes nothing, so it
# is allowed without sign-off (it is the preview used to DECIDE sign-off).
SIGNOFF_RE = re.compile(r"^\s*-\s*\[[xX]\]\s*checkpoint\s*:?\s*approved\s+for\s+export",
                        re.MULTILINE | re.IGNORECASE)
UNCHECKED_RE = re.compile(r"^\s*-\s*\[\s*\]\s*(.+)$", re.MULTILINE)


def enforce_checkpoint(review_path, dry_run):
    if dry_run:
        print("Checkpoint: --dry-run (nothing written) — gate not enforced.")
        return
    if not review_path:
        fail("checkpoint not signed off: a real export requires --review "
             "<slug>_review.md with a human-checked sign-off line. Use --dry-run "
             "to preview without sign-off.")
    p = Path(review_path)
    if not p.exists():
        fail(f"review note not found: {p}")
    text = p.read_text(encoding="utf-8")
    unresolved = [m.group(1).strip() for m in UNCHECKED_RE.finditer(text)]
    if not SIGNOFF_RE.search(text):
        msg = (f"CHECKPOINT NOT SIGNED OFF — export blocked.\n"
               f"  {p.name} has no checked sign-off line. After resolving every\n"
               f"  open item, add this exact line to the note:\n"
               f"    - [x] Checkpoint: approved for export")
        if unresolved:
            msg += "\n  Unresolved items still open:\n" + "\n".join(
                f"    - {u}" for u in unresolved)
        fail(msg)
    if unresolved:
        fail("CHECKPOINT inconsistent — export blocked: a sign-off line is present "
             "but unresolved items remain. Resolve (check) them first:\n"
             + "\n".join(f"    - {u}" for u in unresolved))
    print(f"Checkpoint: signed off in {p.name} — proceeding with export.")


def find_sheet(wb, target):
    """Match a sheet by exact name, then case-insensitively. Fail loud."""
    if target in wb.sheetnames:
        return wb[target]
    lower = {n.lower(): n for n in wb.sheetnames}
    if target.lower() in lower:
        return wb[lower[target.lower()]]
    fail(f"sheet {target!r} not found. Available: {wb.sheetnames}")


def last_data_row(ws, key_col):
    """Last row (1-indexed) with a non-empty value in key_col; 0 if none."""
    last = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last


def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def to_int(cell):
    s = (cell or "").strip().replace(",", "")
    if s == "":
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        fail(f"non-numeric value in a numeric column: {cell!r}")


def fill_rgb(cell):
    """Solid-fill ARGB hex for a cell, or None if unfilled."""
    f = cell.fill
    return f.fgColor.rgb if (f and f.patternType) else None


def solid(color):
    return PatternFill("solid", fgColor=color)


def append_granular(ws, rows, dry_run):
    if not rows:
        return 0, 0, 0
    header, data = rows[0], rows[1:]
    if len(header) != GRANULAR_NCOLS:
        fail(f"granular CSV has {len(header)} columns, expected {GRANULAR_NCOLS} (A-P)")
    start = last_data_row(ws, 1) + 1  # Source col A always populated
    n_uncertain = n_red = 0
    for i, row in enumerate(data):
        if len(row) != GRANULAR_NCOLS:
            fail(f"granular row {i+1} has {len(row)} cols, expected {GRANULAR_NCOLS}")
        r = start + i
        note = (row[15] or "").lower()
        is_uncertain = "flag:uncertain" in note
        is_red = "flag:red" in note
        # continue the zebra banding: even sheet row -> light blue, odd -> white
        band = GRANULAR_BANDS[0] if r % 2 == 0 else GRANULAR_BANDS[1]
        if not dry_run:
            ws.row_dimensions[r].height = GRANULAR_ROW_H
        for col1 in range(1, GRANULAR_NCOLS + 1):
            raw = row[col1 - 1]
            is_num = col1 in NUMERIC_COLS
            val = to_int(raw) if is_num else (raw or None)
            # decide the flag fill for this cell: red on Total wins (known-wrong),
            # then uncertain on any populated numeric cell, else the zebra band.
            flag_fill = None
            if is_num and val is not None:
                if is_red and col1 == TOTAL_COL:
                    flag_fill = RED_FILL
                    n_red += 1
                elif is_uncertain:
                    flag_fill = UNCERTAIN_FILL
                    n_uncertain += 1
            if dry_run:
                continue
            cell = ws.cell(row=r, column=col1, value=val)
            cell.font = copy(GRANULAR_FONT)
            cell.alignment = copy(GRANULAR_ALIGN)
            cell.fill = solid(flag_fill) if flag_fill else solid(band)
            if is_num:
                cell.number_format = NUM_FORMAT
    return len(data), n_uncertain, n_red


def append_sheet1(ws, rows, link_data, dry_run):
    if len(rows) < 2:
        fail("sheet1 CSV has no data row")
    fields = rows[1]
    if len(fields) != SHEET1_NFIELDS:
        fail(f"sheet1 CSV has {len(fields)} fields, expected {SHEET1_NFIELDS} (B-I)")
    ref = last_data_row(ws, SHEET1_START_COL)  # last existing data row (col B)
    start = ref + 1  # Source Overview = col B
    if not dry_run:
        for j, val in enumerate(fields):
            col = SHEET1_START_COL + j
            cell = ws.cell(row=start, column=col, value=(val or None))
            # clone the existing data-row style so new rows match even past the
            # workbook's pre-formatted zone (borders, font, wrap, alignment)
            src = ws.cell(row=ref, column=col)
            cell.font = copy(src.font)
            cell.border = copy(src.border)
            cell.alignment = copy(src.alignment)
            cell.fill = copy(src.fill)
            cell.number_format = src.number_format
        if link_data:
            ws.cell(row=start, column=SHEET1_LINK_DATA_COL, value=link_data)
        if ws.row_dimensions[ref].height:
            ws.row_dimensions[start].height = ws.row_dimensions[ref].height
    return start


def parse_alias_groups(ws):
    """Parse the TA-grouped Aliases tab. Each group = a header row (col A set,
    col B empty, bold) followed by its data rows (col B = TA). Returns a list of
    {name, header, data:[rows], fill} in sheet order."""
    groups = []
    cur = None
    for r in range(3, ws.max_row + 1):  # rows 1-2 are column header + HOW-TO
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        a_txt = str(a).strip() if a not in (None, "") else ""
        if a_txt and (b in (None, "")) and ws.cell(row=r, column=1).font.bold:
            cur = {"name": a_txt, "header": r, "data": [], "fill": None}
            groups.append(cur)
        elif b not in (None, "") and cur is not None:
            cur["data"].append(r)
            if cur["fill"] is None:
                cur["fill"] = fill_rgb(ws.cell(row=r, column=1))
    return groups


def normalize_alias_heights(ws):
    """openpyxl's insert_rows shifts cells but NOT row-height definitions, so
    after inserts the heights drift out of alignment with header/data rows.
    Re-assert: header rows = ALIAS_HEADER_H, data rows = ALIAS_ROW_H."""
    for g in parse_alias_groups(ws):
        ws.row_dimensions[g["header"]].height = ALIAS_HEADER_H
        for dr in g["data"]:
            ws.row_dimensions[dr].height = ALIAS_ROW_H


def style_alias_data_row(ws, r, fill):
    for c in (1, 2, 3):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = solid(fill)
    ws.row_dimensions[r].height = ALIAS_ROW_H


def insert_aliases_sorted(ws, rows, dry_run):
    """Insert each new alias into its TA group at the correct alphabetical slot
    (by Canonical Indication Name, col A), styled to match the group. Returns
    (count_inserted, plan_lines, warnings)."""
    if not rows or len(rows) < 2:
        return 0, [], []
    data = [r for r in rows[1:] if any((c or "").strip() for c in r)]
    plan, warnings, n = [], [], 0
    for row in data:
        canonical = (row[0] or "").strip()
        ta = (row[1] or "").strip()
        aliases = (row[2] or "").strip() if len(row) > 2 else ""
        if not canonical or not ta:
            fail(f"alias row missing canonical name or TA: {row!r}")
        groups = parse_alias_groups(ws)  # re-parse each pass (rows shift on insert)
        match = next((g for g in groups if g["name"].lower() == ta.lower()), None)

        if match is None:
            # No such TA group — create one at the bottom and flag for review.
            hdr = ws.max_row + 1
            if not dry_run:
                ref_hdr = groups[0]["header"] if groups else None
                hc = ws.cell(row=hdr, column=1, value=ta)
                if ref_hdr:
                    src = ws.cell(row=ref_hdr, column=1)
                    hc.font = copy(src.font)
                    hc.fill = copy(src.fill)
                    hc.alignment = copy(src.alignment)
                else:
                    hc.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
                    hc.fill = solid(ALIAS_HEADER_FILL)
                ws.row_dimensions[hdr].height = ALIAS_HEADER_H
                drow = hdr + 1
                ws.cell(row=drow, column=1, value=canonical)
                ws.cell(row=drow, column=2, value=ta)
                ws.cell(row=drow, column=3, value=aliases or None)
                style_alias_data_row(ws, drow, ALIAS_NEWGROUP_FILL)
            plan.append(f"  {canonical!r} -> NEW group {ta!r} (header row {hdr})")
            warnings.append(f"  WARNING: TA group {ta!r} did not exist; created at "
                            f"row {hdr} with placeholder color {ALIAS_NEWGROUP_FILL}. "
                            f"Verify group position/color manually.")
            n += 1
            continue

        # Find alphabetical insertion point within the group (by col A).
        at = match["data"][-1] + 1 if match["data"] else match["header"] + 1
        for dr in match["data"]:
            existing = str(ws.cell(row=dr, column=1).value or "")
            if canonical.lower() < existing.strip().lower():
                at = dr
                break
        if not dry_run:
            ws.insert_rows(at)
            ws.cell(row=at, column=1, value=canonical)
            ws.cell(row=at, column=2, value=ta)
            ws.cell(row=at, column=3, value=aliases or None)
            style_alias_data_row(ws, at, match["fill"])
        plan.append(f"  {canonical!r} -> group {ta!r} at row {at} (fill {match['fill']})")
        n += 1
    if n and not dry_run:
        normalize_alias_heights(ws)
    return n, plan, warnings


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--granular", required=True)
    ap.add_argument("--sheet1", required=True)
    ap.add_argument("--aliases")
    ap.add_argument("--review", help="path to <slug>_review.md; a real (non-dry-run) "
                    "export is blocked unless it carries a checked sign-off line")
    ap.add_argument("--link-data")
    ap.add_argument("--out")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    # HARD checkpoint gate — must pass before any write happens.
    enforce_checkpoint(args.review, args.dry_run)

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        fail(f"workbook not found: {wb_path}")
    out_path = Path(args.out) if args.out else wb_path.with_name(wb_path.stem + "_UPDATED.xlsx")

    wb = load_workbook(wb_path)  # keep formulas + all tabs intact
    g_ws = find_sheet(wb, GRANULAR_SHEET)
    s_ws = find_sheet(wb, SHEET1_SHEET)

    g_rows = read_csv(args.granular)
    s_rows = read_csv(args.sheet1)

    g_start = last_data_row(g_ws, 1) + 1
    n_gran, n_uncertain, n_red = append_granular(g_ws, g_rows, args.dry_run)
    s_start = append_sheet1(s_ws, s_rows, args.link_data, args.dry_run)

    n_alias = 0
    alias_plan, alias_warnings = [], []
    if args.aliases:
        a_ws = find_sheet(wb, ALIASES_SHEET)
        n_alias, alias_plan, alias_warnings = insert_aliases_sorted(
            a_ws, read_csv(args.aliases), args.dry_run)

    print("=== pharma-epi-exporter ===")
    print(f"workbook : {wb_path.name}")
    print(f"Granular Data : {n_gran} row(s) -> rows {g_start}..{g_start + n_gran - 1}"
          if n_gran else "Granular Data : 0 rows")
    print(f"  flag fills: {n_uncertain} uncertain (FFD966), {n_red} red (Total known-wrong)")
    print(f"Sheet1 : 1 summary row -> row {s_start} (data from col B)")
    if not args.link_data:
        print("  NOTE: Link-to-data column (G) left blank — pass --link-data to fill.")
    if args.aliases:
        print(f"Indication Aliases : {n_alias} row(s) inserted into TA groups")
        for line in alias_plan:
            print(line)
        for line in alias_warnings:
            print(line)

    if args.dry_run:
        print("\nDRY RUN — no file written.")
        return
    wb.save(out_path)
    print(f"\nWrote: {out_path}")
    print("Review it, then upload to Box as a new version of the master sheet.")


if __name__ == "__main__":
    main()
